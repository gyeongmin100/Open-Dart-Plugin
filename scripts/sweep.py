# -*- coding: utf-8 -*-
"""경우의 수 스윕 — 실제 DART 공시로 문서 선택·생성·검증을 끝까지 돌린다.

pytest가 아니다. DART_API_KEY와 네트워크가 필요하고 수백 MB를 내려받는다.

    python scripts/sweep.py --per-axis 10
    python scripts/sweep.py --rcept 20250814004433   # 실패 건만 재실행

실패는 네 갈래로 분류한다.
  structural  — 첨부도 본문 재무제표도 애초에 없음 (정상)
  source      — 원문 자체의 오류 (경고로 처리되는 것)
  api         — DART 쪽 오류·연결 끊김 (동시 요청을 줄이면 준다)
  bug         — 나머지. 0이어야 한다.

DART 화면(dsaf001)은 동시 요청에 민감하다. --concurrency를 올리면
api 실패가 늘어난다.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from opendartmcp.client import DartClient          # noqa: E402
import httpx                                        # noqa: E402

from opendartmcp.errors import DartApiError         # noqa: E402
from opendartmcp.excel import candidates, workflow  # noqa: E402

# (축 이름, list.json 파라미터) — 기간·시장·외부감사 단독을 교차한다.
AXES = [
    ("연간·유가", {"pblntf_detail_ty": "A001", "corp_cls": "Y"}),
    ("연간·코스닥", {"pblntf_detail_ty": "A001", "corp_cls": "K"}),
    ("반기·유가", {"pblntf_detail_ty": "A002", "corp_cls": "Y"}),
    ("반기·코스닥", {"pblntf_detail_ty": "A002", "corp_cls": "K"}),
    ("분기·유가", {"pblntf_detail_ty": "A003", "corp_cls": "Y"}),
    ("분기·코스닥", {"pblntf_detail_ty": "A003", "corp_cls": "K"}),
    ("분기·코넥스", {"pblntf_detail_ty": "A003", "corp_cls": "N"}),
    ("외감단독·비상장", {"pblntf_ty": "F", "corp_cls": "E"}),
]
# 정정은 제목으로만 구분된다.
CORRECTION_AXIS = ("정정", {"pblntf_ty": "A"})
CORRECTION_MARKS = ("정정", "추가")

PERIODS = {"A001": ("0301", "0331"), "A002": ("0801", "0831"),
           "A003": ("1101", "1130")}
DEFAULT_PERIOD = ("0401", "0430")

EXCLUDE = ("감사의감사보고서", "내부회계관리제도운영보고서",
           "내부감시장치에대한감사의의견서", "영업보고서", "정관",
           "자기주식보고서", "기업개황자료")
SOURCE_ERROR_MARKS = ("원문에 없는 주석번호",)


def _norm(text: str) -> str:
    return text.replace(" ", "")


def choose(candidate_list: list[dict], scope: str) -> tuple[dict | None, str]:
    """SKILL의 선택 규칙을 코드로 옮긴 것. (후보, 종류)를 반환한다."""
    audit = [c for c in candidate_list
             if any(k in _norm(c["title"]) for k in ("감사보고서", "검토보고서"))
             and not any(x in _norm(c["title"]) for x in EXCLUDE)]
    wanted = scope == "consolidated"
    match = [c for c in audit if ("연결" in _norm(c["title"])) == wanted]
    if match:
        return max(match, key=lambda c: c["date"]), "attachment"
    body = [c for c in candidate_list if c["title"].endswith(candidates.BODY_SUFFIX)]
    if body:
        return body[0], "body"
    return None, ""


def classify(kind: str, result: dict) -> str:
    """실패 원인 분류 — bug가 아닌 것만 좁게 인정한다."""
    error = result.get("error")
    if error in ("scope_not_in_document", "scope_not_applicable"):
        return "structural"      # 그 범위 재무제표가 문서에 없다
    if error == "verification_failed":
        if any(mark in " ".join(result.get("failures", []))
               for mark in SOURCE_ERROR_MARKS):
            return "source"
    return "bug"


def has_opposite_scope(candidate_list: list[dict], scope: str) -> bool:
    """요청 범위 후보는 없지만 반대 범위 감사·검토보고서는 있는지."""
    audit = [c for c in candidate_list
             if any(k in _norm(c["title"]) for k in ("감사보고서", "검토보고서"))
             and not any(x in _norm(c["title"]) for x in EXCLUDE)]
    wanted = scope == "consolidated"
    return bool(audit) and all(("연결" in _norm(c["title"])) != wanted
                               for c in audit)


async def run_one(client, semaphore, axis: str, item: dict, scope: str) -> dict:
    rcept_no, name = item["rcept_no"], item["corp_name"]
    row = {"axis": axis, "rcept_no": rcept_no, "corp": name, "scope": scope,
           "report_nm": item.get("report_nm", "")}
    async with semaphore:
        try:
            listed = await candidates.list_candidates(client, rcept_no)
        except DartApiError as error:                    # DART 쪽 사정이다
            return {**row, "status": "api", "detail": f"list:{error.status}"}
        except httpx.TransportError as error:            # 연결 끊김도 마찬가지
            return {**row, "status": "api", "detail": f"list:{type(error).__name__}"}
        except Exception as error:                       # noqa: BLE001
            return {**row, "status": "bug", "detail": f"list:{type(error).__name__}"}
        if not listed.get("ok"):
            error = listed.get("error")
            if error == "confirmation_required":
                return {**row, "status": "confirmation_required",
                        "detail": "본문 사용 전 사용자 승인 필요"}
            if error == "attachment_lookup_failed":
                return {**row, "status": "api", "detail": error}
            return {**row, "status": "bug", "detail": error}
        chosen, kind = choose(listed["candidates"], scope)
        if chosen is None:
            if has_opposite_scope(listed["candidates"], scope):
                return {**row, "status": "not_applicable",
                        "detail": "반대 범위 문서만 존재"}
            return {**row, "status": "structural", "detail": "후보 없음"}
        row["candidate"] = chosen["title"]
        row["kind"] = kind
        try:
            result = await workflow.create_workbook(
                client, candidate_id=chosen["candidate_id"], scope=scope,
                output_dir=tempfile.mkdtemp(), output_name="sweep.xlsx")
        except DartApiError as error:
            return {**row, "status": "api", "detail": f"build:{error.status}"}
        except workflow.StageFailure as error:
            status = "api" if error.stage == "download" else "bug"
            return {**row, "status": status,
                    "detail": f"{error.stage}:{error.error_type}"}
        except Exception as error:                       # noqa: BLE001
            return {**row, "status": "bug", "detail": f"raise:{type(error).__name__}"}
    if result.get("ok"):
        verification = result["verification"]
        # 어떤 재무제표 종류를 실제로 밟았는지 남긴다 — 일반기업회계기준
        # (대차대조표·이익잉여금처분계산서) 커버 여부가 여기서 드러난다.
        from openpyxl import load_workbook
        sheets = [s for s in load_workbook(result["workbook"]).sheetnames
                  if s != "주석"]
        return {**row, "status": "ok", "sheets": sheets,
                "detail": (f"표{verification['statements']} "
                           f"주석{verification['notes']} "
                           f"링크{verification['hyperlinks']}"),
                "warnings": verification["warnings"]}
    return {**row, "status": classify(kind, result),
            "detail": json.dumps(result, ensure_ascii=False)[:240]}


async def sample(client, axis: str, params: dict, count: int, year: str,
                 corrections_only: bool = False) -> list[dict]:
    begin, end = PERIODS.get(params.get("pblntf_detail_ty", ""), DEFAULT_PERIOD)
    try:
        response = await client.get_json("/list.json", {
            "bgn_de": year + begin, "end_de": year + end,
            "page_count": "100", **params})
    except DartApiError as error:
        if error.status != "013":       # 그 해에 그 시장이 없었을 뿐이다
            raise
        return []
    items = response.get("list", []) or []
    if corrections_only:
        items = [i for i in items
                 if any(mark in i["report_nm"] for mark in CORRECTION_MARKS)]
    # 표집은 재현 가능해야 한다 — 무작위 대신 균등 간격으로 고른다.
    if len(items) > count:
        step = len(items) / count
        items = [items[int(i * step)] for i in range(count)]
    return [(axis, item) for item in items]


async def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--per-axis", type=int, default=6)
    parser.add_argument("--concurrency", type=int, default=3)
    parser.add_argument("--year", action="append", default=[],
                        help="표집 연도(여러 번 지정 가능). 기본 2025")
    parser.add_argument("--rcept", action="append", default=[],
                        help="이 접수번호만 돌린다(실패 재현용)")
    parser.add_argument("--out", default="sweep_result.json")
    args = parser.parse_args()

    client = DartClient(os.environ["DART_API_KEY"])
    semaphore = asyncio.Semaphore(args.concurrency)
    try:
        if args.rcept:
            targets = [("재현", {"rcept_no": no, "corp_name": "?"})
                       for no in args.rcept]
        else:
            years = args.year or ["2025"]
            batches = await asyncio.gather(*[
                sample(client, f"{axis}·{year}", params, args.per_axis, year)
                for year in years for axis, params in AXES
            ], *[sample(client, f"{CORRECTION_AXIS[0]}·{year}",
                        CORRECTION_AXIS[1], args.per_axis, year,
                        corrections_only=True) for year in years])
            targets = [pair for batch in batches for pair in batch]

        rows = await asyncio.gather(*[
            run_one(client, semaphore, axis, item, scope)
            for axis, item in targets
            for scope in ("consolidated", "separate")
        ])
    finally:
        await client.aclose()

    tally: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        tally[row["axis"]][row["status"]] += 1

    print(f"{'축':16s} {'OK':>4} {'구조':>4} {'원문':>4} {'API':>4} {'버그':>4}")
    for axis in sorted(tally):
        counts = tally[axis]
        print(f"{axis:16s} {counts['ok']:4d} {counts['structural']:4d} "
              f"{counts['source']:4d} {counts['api']:4d} {counts['bug']:4d}")

    # 축 커버 증명 — 표집이 실제로 그 경우를 밟았는지 보여준다.
    kinds = defaultdict(int)
    for row in rows:
        for sheet in row.get("sheets", []):
            for body in ("대차대조표", "이익잉여금처분계산서", "결손금처리계산서",
                         "포괄손익계산서", "손익계산서", "재무상태표",
                         "자본변동표", "현금흐름표"):
                if body in sheet:
                    kinds[body] += 1
                    break
    print("\n재무제표 종류:", dict(kinds))
    # 사업보고서의 기준월이 곧 결산월이다(반기·분기는 중간 시점이라 제외).
    months = defaultdict(int)
    for row in rows:
        if "사업보고서" not in row.get("report_nm", ""):
            continue
        match = re.search(r"\((\d{4})\.(\d{2})\)", row["report_nm"])
        months[match.group(2) if match else "?"] += 1
    print("결산월(사업보고서 기준):", dict(months))

    bugs = [r for r in rows if r["status"] == "bug"]
    totals = defaultdict(int)
    for row in rows:
        totals[row["status"]] += 1
    print("\n결과:", dict(totals))
    for row in bugs:
        print(f"BUG {row['axis']} {row['rcept_no']} {row['corp']} "
              f"{row['scope']} {row.get('candidate', '')} :: {row['detail']}")
    Path(args.out).write_text(
        json.dumps(rows, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\n총 {len(rows)}건, 버그 {len(bugs)}건 → {args.out}")
    return 1 if bugs else 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
