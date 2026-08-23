import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from opendartmcp.excel.build_financial_excel import build_workbook


def _cell(text, *, header=False, size=None, align=None):
    return {
        "text": text, "align": align, "valign": None,
        "bold": header, "size": size, "font": None,
        "fill": None, "header": header,
    }


def _table(rows, *, note_col=None, borderless=False, col_widths=None):
    return {
        "borderless": borderless,
        "col_widths": col_widths or [],
        "note_col": note_col,
        "rows": rows,
        "merges": [],
    }


def _model(*, note_blocks=None, statement_table=None):
    notes = [{
        "number": 1,
        "blocks": note_blocks or [
            {"type": "paragraph", "text": "1. 주석 제목"},
            {"type": "paragraph", "text": "주석 본문"},
        ],
    }]
    statements = []
    if statement_table:
        statements.append({
            "sheet_name": "재무상태표",
            "preamble": [[_cell("재무상태표")]],
            "tables": [statement_table],
            "postscript": [],
        })
    return {"notes_preamble": [], "notes": notes, "statements": statements}


def _build(model):
    path = Path(tempfile.mkdtemp()) / "layout.xlsx"
    build_workbook(model, str(path))
    return load_workbook(path)


class RequestedLayoutTest(unittest.TestCase):
    def test_note_paragraphs_merge_from_b_through_l(self):
        ws = _build(_model())["주석"]
        paragraph = next(cell for row in ws.iter_rows() for cell in row
                         if cell.value == "주석 본문")
        self.assertIn(f"B{paragraph.row}:L{paragraph.row}",
                      {str(rng) for rng in ws.merged_cells.ranges})

    def test_all_workbook_text_is_one_point_larger(self):
        statement = _table([
            {"header": True, "cells": [
                _cell("과목", header=True), _cell("주석", header=True),
                _cell("당기", header=True),
            ]},
            {"header": False, "cells": [
                _cell("현금", size=9), _cell("1"), _cell("100"),
            ]},
        ], note_col=1)
        wb = _build(_model(statement_table=statement))
        notes = wb["주석"]
        fs = wb["재무상태표"]
        by_value = lambda ws, value: next(
            cell for row in ws.iter_rows() for cell in row if cell.value == value)

        self.assertEqual(by_value(notes, "1. 주석 제목").font.sz, 14)
        self.assertEqual(by_value(notes, "주석 본문").font.sz, 12)
        self.assertEqual(by_value(fs, "재무상태표").font.sz, 15)
        self.assertEqual(by_value(fs, "현금").font.sz, 10)
        self.assertEqual(by_value(fs, 100).font.sz, 12)

    def test_note_paragraph_height_has_room_for_larger_font(self):
        ws = _build(_model())["주석"]
        body = next(cell for row in ws.iter_rows() for cell in row
                    if cell.value == "주석 본문")
        self.assertGreaterEqual(ws.row_dimensions[body.row].height, 22)

    def test_statement_note_columns_are_last_and_grow_proportionally(self):
        statement = _table([
            {"header": True, "cells": [
                _cell("과목", header=True), _cell("주석", header=True),
                _cell("당기", header=True), _cell("전기", header=True),
            ]},
            {"header": False, "cells": [
                _cell("현금"), _cell("1, 2, 3"),
                _cell("100"), _cell("90"),
            ]},
        ], note_col=1, col_widths=[210, 35, 100, 100])
        ws = _build(_model(statement_table=statement))["재무상태표"]
        header_row = next(row for row in ws.iter_rows()
                          if any(cell.value == "과목" for cell in row))
        account = next(cell for cell in header_row if cell.value == "과목")
        current = next(cell for cell in header_row if cell.value == "당기")
        prior = next(cell for cell in header_row if cell.value == "전기")
        note = next(cell for cell in header_row if cell.value == "주석")

        self.assertEqual(current.column, account.column + 1)
        self.assertEqual(prior.column, current.column + 1)
        self.assertEqual(note.column, prior.column + 1)
        note_widths = [ws.column_dimensions[
            get_column_letter(note.column + i)].width for i in range(3)]
        self.assertTrue(all(width >= 3.14 * 12 / 11 for width in note_widths))

    def test_unit_row_spans_table_width_and_is_right_aligned(self):
        unit = _table([
            {"header": False, "cells": [
                _cell("   (단위: 천원)", align="left")
            ]},
        ], borderless=True)
        data = _table([
            {"header": True, "cells": [
                _cell("구분", header=True), _cell("2022", header=True),
                _cell("2021", header=True), _cell("2020", header=True),
            ]},
            {"header": False, "cells": [
                _cell("현금"), _cell("100"), _cell("90"), _cell("80"),
            ]},
        ])
        blocks = [
            {"type": "paragraph", "text": "1. 주석 제목"},
            {"type": "table", "table": unit},
            {"type": "table", "table": data},
        ]
        ws = _build(_model(note_blocks=blocks))["주석"]
        cell = next(cell for row in ws.iter_rows() for cell in row
                    if isinstance(cell.value, str) and "(단위:" in cell.value)

        self.assertIn(f"B{cell.row}:E{cell.row}",
                      {str(rng) for rng in ws.merged_cells.ranges})
        self.assertEqual(cell.alignment.horizontal, "right")


if __name__ == "__main__":
    unittest.main()
