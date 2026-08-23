import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from opendartmcp.excel import dartdoc
from opendartmcp.excel.build_financial_excel import build_workbook
from tests.fixtures import annual_body_xml, audit_report_xml


class ExcelPackageImportTest(unittest.TestCase):
    def test_package_imports_with_relative_imports(self):
        from opendartmcp.excel import build_financial_excel, dartdoc, verify_workbook

        self.assertTrue(callable(build_financial_excel.build_workbook))
        self.assertTrue(callable(verify_workbook.verify))
        self.assertEqual(dartdoc.CONSOLIDATED, "consolidated")

    def test_no_sys_path_or_ensure_deps(self):
        from pathlib import Path

        import opendartmcp.excel as pkg

        for path in Path(pkg.__file__).parent.glob("*.py"):
            source = path.read_text(encoding="utf-8")
            self.assertNotIn("sys.path.insert", source, path.name)
            self.assertNotIn("_ensure_deps", source, path.name)


class StatementNoteColumnTest(unittest.TestCase):
    def test_note_column_keeps_source_position_and_compact_number_cells(self):
        path = Path(tempfile.mkdtemp()) / "notes.xlsx"
        model = dartdoc.extract_model(
            audit_report_xml("consolidated"), dartdoc.CONSOLIDATED)
        table = model["statements"][0]["tables"][0]
        table["rows"][1]["cells"][table["note_col"]]["text"] = \
            "1, 2, 123456789"
        build_workbook(model, str(path))
        ws = load_workbook(path)[model["statements"][0]["sheet_name"]]

        header_row = next(row for row in ws.iter_rows()
                          if any(cell.value == "과목" for cell in row))
        header = [cell.value for cell in header_row]
        account_col = header.index("과목") + 1
        note_col = header.index("주석") + 1
        amount_col = header.index("당기") + 1
        self.assertEqual(note_col, account_col + 1)
        self.assertEqual(amount_col, note_col + 3)

        data_row = header_row[0].row + 1
        note_cells = [ws.cell(data_row, note_col + i) for i in range(3)]
        self.assertEqual([cell.value for cell in note_cells], [1, 2, 123456789])
        self.assertTrue(all(isinstance(cell.value, int) for cell in note_cells))
        self.assertIsNone(note_cells[0].border.right.style)
        self.assertIsNone(note_cells[1].border.left.style)
        self.assertIsNone(note_cells[1].border.right.style)
        self.assertIsNone(note_cells[2].border.left.style)
        self.assertEqual(note_cells[2].border.right.style, "thin")
        widths = [ws.column_dimensions[cell.column_letter].width
                  for cell in note_cells]
        self.assertEqual(widths[:2], [3.5, 3.5])
        self.assertGreater(widths[2], 8)

    def test_missing_source_note_column_stays_omitted(self):
        path = Path(tempfile.mkdtemp()) / "body.xlsx"
        model = dartdoc.extract_model(
            annual_body_xml("consolidated"), dartdoc.CONSOLIDATED)
        build_workbook(model, str(path))
        wb = load_workbook(path)
        ws = wb[model["statements"][0]["sheet_name"]]
        self.assertNotIn("주석", [cell.value for row in ws.iter_rows()
                                  for cell in row])
