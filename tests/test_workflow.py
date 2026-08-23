import json
import re
import tempfile
import unittest
from pathlib import Path

import httpx
from openpyxl import load_workbook

from opendartmcp.excel import candidates, dartdoc, workflow
from tests.fixtures import (annual_body_xml, annual_report_zip, audit_report_xml,
                            body_only_zip, make_zip, quarterly_report_zip,
                            titled_audit_report_xml, viewer_fragment, viewer_page)
from tests.test_client_zip_helpers import make_client

RCEPT = "20250319000665"
BODY_ID = f"body:{RCEPT}"


class WorkflowTestBase(unittest.IsolatedAsyncioTestCase):
    def setUp(self):
        self.out = Path(tempfile.mkdtemp())
        self.calls = []

    async def run_workflow(self, zip_bytes, **kwargs):
        client = make_client(lambda r: httpx.Response(200, content=zip_bytes),
                             count=self.calls)
        params = {"candidate_id": BODY_ID, "scope": "consolidated",
                  "output_dir": str(self.out), "output_name": "결과.xlsx",
                  "allow_body": True}
        params.update(kwargs)
        try:
            return await workflow.create_workbook(client, **params)
        finally:
            await client.aclose()


class BodyDocumentTest(WorkflowTestBase):
    async def test_body_candidate_uses_periodic_report_not_first_entry(self):
        """ZIP에 첨부가 먼저 와도 본문은 정기보고서다."""
        zip_bytes = make_zip({
            "000_attachment.xml": audit_report_xml(
                "separate", title="감사보고서"),
            "999_body.xml": annual_body_xml("consolidated"),
        })
        result = await self.run_workflow(zip_bytes)
        self.assertTrue(result["ok"], result)
        self.assertEqual(result["source_title"], "사업보고서")

    async def test_body_of_periodic_report_allows_zero_links(self):
        result = await self.run_workflow(body_only_zip(RCEPT))
        self.assertTrue(result["ok"], result)
        self.assertTrue(result["used_body"])
        self.assertEqual(result["verification"]["hyperlinks"], 0)

    async def test_standalone_audit_filing_is_not_treated_as_body(self):
        """외부감사 단독 공시는 ZIP의 유일한 문서가 감사보고서다.

        본문 후보로 골라도 첨부 형식이므로 링크 검사를 풀면 안 된다.
        """
        zip_bytes = make_zip({f"{RCEPT}_00760.xml": audit_report_xml(
            "separate", title="감사보고서")})
        result = await self.run_workflow(zip_bytes, scope="separate")
        self.assertTrue(result["ok"], result)
        self.assertFalse(result["used_body"])
        self.assertGreater(result["verification"]["hyperlinks"], 0)

    async def test_title_tag_statements_are_extracted(self):
        """비상장 외감법인 형식: 재무제표 제목이 TITLE 태그다."""
        zip_bytes = make_zip({f"{RCEPT}_00760.xml": titled_audit_report_xml()})
        result = await self.run_workflow(zip_bytes, scope="separate")
        self.assertTrue(result["ok"], result)
        self.assertEqual(result["verification"]["statements"], 4)
        self.assertGreater(result["verification"]["hyperlinks"], 0)

    async def test_downloads_document_xml_exactly_once(self):
        await self.run_workflow(annual_report_zip(RCEPT))
        self.assertEqual(self.calls, ["/api/document.xml"])

    async def test_server_does_not_look_up_other_rcept_no(self):
        await self.run_workflow(body_only_zip(RCEPT))
        self.assertEqual(len(self.calls), 1)


class ErrorMappingTest(WorkflowTestBase):
    async def test_invalid_candidate_id_does_not_download(self):
        result = await self.run_workflow(annual_report_zip(RCEPT),
                                         candidate_id="123")
        self.assertEqual(result["error"], "invalid_candidate_id")
        self.assertEqual(self.calls, [])

    async def test_invalid_output_dir(self):
        result = await self.run_workflow(annual_report_zip(RCEPT),
                                         output_dir=str(self.out / "missing"))
        self.assertEqual(result["error"], "invalid_output_dir")
        self.assertEqual(self.calls, [])

    async def test_output_name_with_path_separator_is_rejected(self):
        result = await self.run_workflow(annual_report_zip(RCEPT),
                                         output_name="../escape.xlsx")
        self.assertEqual(result["error"], "invalid_output_dir")
        self.assertEqual(self.calls, [])

    async def test_scope_not_in_document_returns_available_scopes(self):
        """§5.6: 별도 문서에 연결을 요청하면 있는 범위를 알려준다."""
        zip_bytes = make_zip({f"{RCEPT}_1.xml": audit_report_xml(
            "separate", title="연결감사보고서")})
        result = await self.run_workflow(zip_bytes)
        self.assertEqual(result["error"], "scope_not_in_document")
        self.assertEqual(result["available_scopes"], ["separate"])

    async def test_no_financial_statements_has_distinct_code(self):
        """표는 있으나 제목을 못 읽은 경우 — 재시도가 무의미한 실패."""
        zip_bytes = make_zip({f"{RCEPT}_1.xml": re.sub(
            r"<TR><TD>연결[가-힣]+</TD></TR>", "<TR><TD>내역</TD></TR>",
            audit_report_xml("consolidated", title="연결감사보고서"))})
        result = await self.run_workflow(zip_bytes)
        self.assertEqual(result["error"], "no_financial_statements")
        self.assertEqual(result["reason"], "no_statement_tables")

    async def test_scope_not_applicable_when_section_is_empty(self):
        """종속회사가 없어 연결재무제표를 안 싣는 경우는 별개 코드다."""
        zip_bytes = make_zip({f"{RCEPT}_1.xml": audit_report_xml(
            "consolidated", with_tables=False, title="연결감사보고서")})
        result = await self.run_workflow(zip_bytes)
        self.assertEqual(result["error"], "scope_not_applicable")

    async def test_empty_zip_is_candidate_unavailable(self):
        result = await self.run_workflow(make_zip({}))
        self.assertEqual(result["error"], "candidate_unavailable")

    async def test_library_errors_do_not_exit_or_print(self):
        import contextlib
        import io as _io

        buffer = _io.StringIO()
        with contextlib.redirect_stdout(buffer):
            result = await self.run_workflow(make_zip({}))
        self.assertFalse(result["ok"])
        self.assertEqual(buffer.getvalue(), "")


class OutputFileTest(WorkflowTestBase):
    async def test_existing_file_is_preserved_and_suffixed(self):
        (self.out / "결과.xlsx").write_bytes(b"old")
        result = await self.run_workflow(annual_report_zip(RCEPT))
        self.assertEqual((self.out / "결과.xlsx").read_bytes(), b"old")
        self.assertEqual(Path(result["workbook"]).name, "결과 (1).xlsx")

    async def test_no_temp_files_left_on_success(self):
        result = await self.run_workflow(annual_report_zip(RCEPT))
        self.assertEqual([p.name for p in self.out.iterdir()],
                         [Path(result["workbook"]).name])

    async def test_failed_verification_leaves_no_workbook(self):
        original = workflow.verify
        workflow.verify = lambda *a, **k: {
            "ok": False, "failures": ["강제 실패"], "warnings": [], "stats": {}}
        try:
            result = await self.run_workflow(annual_report_zip(RCEPT))
        finally:
            workflow.verify = original
        self.assertEqual(result["error"], "verification_failed")
        self.assertEqual(result["failures"], ["강제 실패"])
        self.assertEqual(list(self.out.iterdir()), [])

    async def test_cancellation_cleans_up_temp_file(self):
        original = workflow.build_workbook

        def boom(model, path):
            Path(path).write_bytes(b"partial")
            raise KeyboardInterrupt()

        workflow.build_workbook = boom
        try:
            with self.assertRaises(KeyboardInterrupt):
                await self.run_workflow(annual_report_zip(RCEPT))
        finally:
            workflow.build_workbook = original
        self.assertEqual(list(self.out.iterdir()), [])

    async def test_final_move_failure_cleans_reserved_file(self):
        original = workflow.os.replace

        def boom(source, destination):
            raise OSError("move failed")

        workflow.os.replace = boom
        try:
            with self.assertRaises(workflow.StageFailure):
                await self.run_workflow(annual_report_zip(RCEPT))
        finally:
            workflow.os.replace = original
        self.assertEqual(list(self.out.iterdir()), [])

    async def test_zip_entry_path_is_not_used_as_file_path(self):
        """Zip Slip 방지: 엔트리 경로로 파일을 쓰지 않는다 (plan.md §8)."""
        zip_bytes = make_zip({
            "../../evil_00761.xml": audit_report_xml("consolidated",
                                                     title="연결감사보고서")})
        result = await self.run_workflow(zip_bytes)
        self.assertTrue(result["ok"], result)
        self.assertEqual(Path(result["workbook"]).parent, self.out)
        self.assertFalse((self.out.parent.parent / "evil_00761.xml").exists())


class ResultSizeTest(WorkflowTestBase):
    async def test_result_has_no_raw_content(self):
        result = await self.run_workflow(annual_report_zip(RCEPT))
        blob = json.dumps(result, ensure_ascii=False)
        for forbidden in ("<DOCUMENT", "<TABLE", "현금및현금성자산", "PK\x03\x04"):
            self.assertNotIn(forbidden, blob)
        self.assertNotIn("content", result)
        self.assertLess(len(blob.encode("utf-8")), 4096)

    async def test_error_result_has_no_raw_content(self):
        zip_bytes = make_zip({f"{RCEPT}_1.xml": audit_report_xml(
            "consolidated", with_tables=False, title="연결감사보고서")})
        result = await self.run_workflow(zip_bytes)
        blob = json.dumps(result, ensure_ascii=False)
        self.assertNotIn("<DOCUMENT", blob)
        self.assertLess(len(blob.encode("utf-8")), 4096)

    async def test_success_reports_company_and_verification_stats(self):
        result = await self.run_workflow(annual_report_zip(RCEPT))
        self.assertEqual(result["company"], "테스트 주식회사")
        self.assertEqual(result["rcept_no"], RCEPT)
        self.assertEqual(result["candidate_id"], BODY_ID)
        self.assertEqual(result["scope"], "consolidated")
        self.assertEqual(result["verification"]["statements"], 4)
        self.assertEqual(result["verification"]["notes"], 3)


class QuarterlyTest(WorkflowTestBase):
    async def test_quarterly_body_builds(self):
        result = await self.run_workflow(quarterly_report_zip(RCEPT))
        self.assertTrue(result["ok"], result)
        self.assertTrue(result["used_body"])


class ViewerNotesTest(unittest.IsolatedAsyncioTestCase):
    async def test_zip_statements_keep_html_note_breaks(self):
        dcm = "9085000"
        xml = audit_report_xml("separate", title="감사보고서")
        old_notes = """<TITLE>주석</TITLE>
<P>1. 회사의 개요</P>
<P>당사는 테스트 목적으로 설립되었습니다.</P>
<P>2. 매출채권</P>
<P>매출채권의 내역은 다음과 같습니다.</P>
<P>3. 현금및현금성자산</P>
<P>현금및현금성자산의 내역은 다음과 같습니다.</P>"""
        malformed_notes = """<TITLE>주석</TITLE>
<P>1. 회사의 개요</P>
<P>(1) 첫째 제목설명 문단(2) 둘째 제목</P>
<P>2. 매출채권</P>
<P>매출채권의 내역은 다음과 같습니다.</P>
<P>3. 현금및현금성자산</P>
<P>현금및현금성자산의 내역은 다음과 같습니다.</P>"""
        xml = xml.replace(old_notes, malformed_notes)
        fragment = viewer_fragment().replace(
            '<p class="section-2">주 석</p><p>1. 현금</p>'
            '<p>현금의 내역입니다.</p>',
            '<p class="section-2">주 석</p><p>1. 회사의 개요</p>'
            '<p>(1) 첫째 제목<br><br>설명 문단<br><br>(2) 둘째 제목</p>'
            '<p>2. 매출채권</p><p>매출채권의 내역은 다음과 같습니다.</p>'
            '<p>3. 현금및현금성자산</p>'
            '<p>현금및현금성자산의 내역은 다음과 같습니다.</p>')
        zip_bytes = make_zip({f"{RCEPT}_00760.xml": xml})

        def handler(request):
            if request.url.path == "/api/document.xml":
                return httpx.Response(200, content=zip_bytes)
            if request.url.path == "/dsaf001/main.do":
                return httpx.Response(200, text=viewer_page(
                    RCEPT, [(RCEPT, dcm, "2025.03.19 감사보고서")],
                    node_dcm=dcm, node_text="(첨부)재 무 제 표"))
            if request.url.path == "/report/viewer.do":
                return httpx.Response(200, text=fragment)
            return httpx.Response(404)

        out = Path(tempfile.mkdtemp())
        client = make_client(handler)
        try:
            result = await workflow.create_workbook(
                client, candidate_id=f"zip:{RCEPT}:{RCEPT}_00760.xml",
                scope="separate", output_dir=str(out), output_name="result.xlsx")
        finally:
            await client.aclose()

        self.assertTrue(result["ok"], result)
        ws = load_workbook(result["workbook"])["주석"]
        cells = {str(cell.value): cell for row in ws.iter_rows() for cell in row
                 if cell.value is not None}
        self.assertIn("(1) 첫째 제목", cells)
        self.assertIn("설명 문단", cells)
        self.assertIn("(2) 둘째 제목", cells)
        self.assertNotIn("(1) 첫째 제목설명 문단(2) 둘째 제목", cells)
        self.assertTrue(cells["(1) 첫째 제목"].font.bold)
        self.assertTrue(cells["(2) 둘째 제목"].font.bold)
        first = cells["(1) 첫째 제목"].row
        body = cells["설명 문단"].row
        second = cells["(2) 둘째 제목"].row
        self.assertEqual((body, second), (first + 2, first + 4))
        self.assertIsNone(ws.cell(first + 1, 2).value)
        self.assertIsNone(ws.cell(first + 3, 2).value)
        self.assertEqual(ws.row_dimensions[first + 1].height, 19.2)
        self.assertEqual(ws.row_dimensions[first + 3].height, 19.2)
