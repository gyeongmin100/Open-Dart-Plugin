"""코드 리뷰 지적사항 회귀 테스트 (plan.md §5.11, §5.12, §5.13, §12)."""
import json
import tempfile
import unittest
import zipfile
from pathlib import Path

import httpx
from mcp.server.fastmcp import FastMCP

from opendartmcp.client import DartClient
from opendartmcp.errors import DartApiError
from opendartmcp.excel import dartdoc, workflow
from opendartmcp.excel.build_financial_excel import build_workbook
from opendartmcp.excel.verify_workbook import verify
from opendartmcp.tools import workbook as workbook_tool
from tests.fixtures import (annual_body_xml, annual_report_zip, audit_report_xml,
                            make_zip)
from tests.test_client_zip_helpers import make_client
from tests.test_workbook_tool import _payload

RCEPT = "20250319000665"
API_KEY = "A" * 40


class SecretLeakTest(unittest.IsolatedAsyncioTestCase):
    async def _call(self, handler, **kwargs):
        mcp = FastMCP("test")
        client = DartClient(API_KEY)
        client._http = httpx.AsyncClient(
            base_url=DartClient.BASE_URL, transport=httpx.MockTransport(handler))
        params = {"rcept_no": RCEPT, "scope": "consolidated",
                  "output_dir": ".", "output_name": "x.xlsx"}
        params.update(kwargs)
        try:
            workbook_tool.register(mcp, client)
            result = await mcp.call_tool("create_financial_workbook", params)
        finally:
            await client.aclose()
        return _payload(result)

    async def test_http_error_does_not_leak_api_key(self):
        """§12: 어떤 오류 경로에서도 API 키가 결과에 실려 나가지 않는다."""
        payload = await self._call(lambda r: httpx.Response(500))
        blob = json.dumps(payload, ensure_ascii=False)
        self.assertEqual(payload["error"], "internal_error")
        self.assertNotIn(API_KEY, blob)
        self.assertNotIn("crtfc_key", blob)

    async def test_internal_error_does_not_leak_document_text(self):
        """§5.12: 예기치 못한 예외 메시지로 원문이 흘러나가지 않는다."""
        out = Path(tempfile.mkdtemp())
        secret = "원문에만 있는 문자열"
        original = workflow.build_workbook

        def boom(model, path):
            raise ValueError(f"cell {secret} cannot be used in worksheets")

        workflow.build_workbook = boom
        try:
            payload = await self._call(
                lambda r: httpx.Response(200, content=annual_report_zip(RCEPT)),
                output_dir=str(out))
        finally:
            workflow.build_workbook = original
        self.assertEqual(payload["error"], "internal_error")
        self.assertNotIn(secret, json.dumps(payload, ensure_ascii=False))
        self.assertEqual(list(out.iterdir()), [])

    async def test_internal_error_reports_failing_stage(self):
        """§5.13: 실패 단계를 download/parse/build/verify로 구분한다."""
        payload = await self._call(lambda r: httpx.Response(500))
        self.assertEqual(payload["stage"], "download")

        out = Path(tempfile.mkdtemp())
        original = workflow.build_workbook

        def boom(model, path):
            raise RuntimeError("boom")

        workflow.build_workbook = boom
        try:
            payload = await self._call(
                lambda r: httpx.Response(200, content=annual_report_zip(RCEPT)),
                output_dir=str(out))
        finally:
            workflow.build_workbook = original
        self.assertEqual(payload["stage"], "build")


class LinkCheckScopeTest(unittest.TestCase):
    def _build(self, content):
        model = dartdoc.extract_model(content, dartdoc.CONSOLIDATED)
        path = Path(tempfile.mkdtemp()) / "w.xlsx"
        build_workbook(model, str(path))
        return model, str(path)

    def test_attachment_link_shortage_still_fails(self):
        """첨부 경로의 링크 수 부족은 계속 실패여야 한다 (§12 검증 약화 금지)."""
        content = audit_report_xml("consolidated")
        model, path = self._build(content)
        report = verify(model, path, content, used_body=False)
        self.assertTrue(report["ok"], report["failures"])

        from openpyxl import load_workbook
        wb = load_workbook(path)
        removed = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink is not None and removed < 1:
                        cell.hyperlink = None
                        removed += 1
        wb.save(path)
        report = verify(model, path, content, used_body=False)
        self.assertFalse(report["ok"])
        self.assertTrue(any("주석 링크 수 부족" in f for f in report["failures"]),
                        report["failures"])

    def test_body_path_allows_zero_links(self):
        """§5.7: 본문으로 만든 워크북은 링크 0개가 정상."""
        content = annual_body_xml("consolidated")
        model, path = self._build(content)
        report = verify(model, path, content, used_body=True)
        self.assertTrue(report["ok"], report["failures"])


class OutputPathTest(unittest.IsolatedAsyncioTestCase):
    def setUp(self):
        self.out = Path(tempfile.mkdtemp())

    async def _run(self, **kwargs):
        client = make_client(
            lambda r: httpx.Response(200, content=annual_report_zip(RCEPT)))
        params = {"rcept_no": RCEPT, "scope": "consolidated",
                  "output_dir": str(self.out), "output_name": "결과.xlsx"}
        params.update(kwargs)
        try:
            return await workflow.create_workbook(client, **params)
        finally:
            await client.aclose()

    def test_output_path_is_reserved_not_just_probed(self):
        """§5.11: 이름 선점이 원자적이어야 별도 프로세스가 덮어쓰지 못한다.

        확인 후 rename 방식(`exists()` 검사 + `os.replace`)은 두 프로세스가
        같은 빈 이름을 보고 하나가 다른 하나를 지운다. 선점하면 두 번째 호출은
        곧바로 다른 이름을 받는다.
        """
        first = workflow.reserve_output_path(self.out, "결과.xlsx")
        second = workflow.reserve_output_path(self.out, "결과.xlsx")
        self.assertNotEqual(first, second)
        self.assertEqual(first.name, "결과.xlsx")
        self.assertEqual(second.name, "결과 (1).xlsx")
        self.assertTrue(first.exists())
        self.assertTrue(second.exists())

    async def test_existing_file_is_still_preserved(self):
        """§5.11: 기존 파일 보존 + ` (1)` 규칙은 그대로."""
        (self.out / "결과.xlsx").write_bytes(b"old")
        result = await self._run()
        self.assertEqual((self.out / "결과.xlsx").read_bytes(), b"old")
        self.assertEqual(Path(result["workbook"]).name, "결과 (1).xlsx")

    async def test_unwritable_output_dir_is_invalid_output_dir(self):
        """§5.2/§7: 쓸 수 없는 디렉터리는 internal_error가 아니라 invalid_output_dir.

        Windows에서 os.access(W_OK)는 디렉터리에 대해 사실상 항상 참이므로,
        임시 파일 생성 실패로만 감지할 수 있다.
        """
        import tempfile as _tempfile

        original = _tempfile.mkstemp

        def denied(*args, **kwargs):
            raise PermissionError(13, "Permission denied")

        _tempfile.mkstemp = denied
        try:
            result = await self._run()
        finally:
            _tempfile.mkstemp = original
        self.assertEqual(result["error"], "invalid_output_dir")

    async def test_long_document_title_is_capped(self):
        """§5.12: 비정상적으로 긴 DOCUMENT-NAME이 결과 크기를 키우지 않는다."""
        long_title = "가" * 1500
        zip_bytes = make_zip({
            f"{RCEPT}.xml": annual_body_xml("consolidated").replace(
                "<DOCUMENT-NAME>사업보고서</DOCUMENT-NAME>",
                f"<DOCUMENT-NAME>{long_title}</DOCUMENT-NAME>")})
        client = make_client(lambda r: httpx.Response(200, content=zip_bytes))
        try:
            result = await workflow.create_workbook(
                client, rcept_no=RCEPT, scope="consolidated",
                output_dir=str(self.out), output_name="결과.xlsx")
        finally:
            await client.aclose()
        blob = json.dumps(result, ensure_ascii=False)
        self.assertLess(len(blob.encode("utf-8")), 4096, blob[:200])
        self.assertLessEqual(len(result["documents"][0]["title"]), 200)


class ZipMemberErrorTest(unittest.IsolatedAsyncioTestCase):
    async def test_corrupt_member_maps_to_dart_api_error(self):
        """§5.4: 손상된 엔트리도 기존처럼 DartApiError로 변환된다."""
        good = annual_report_zip(RCEPT)
        corrupt = bytearray(good)
        # 로컬 파일 헤더 뒤 압축 데이터 일부를 훼손해 읽기 시점에 깨지게 한다.
        corrupt[len(corrupt) // 2] ^= 0xFF
        client = make_client(lambda r: httpx.Response(200, content=bytes(corrupt)))
        try:
            with self.assertRaises((DartApiError, zipfile.BadZipFile)) as ctx:
                await client.get_zip_text("/document.xml", {"rcept_no": RCEPT})
        finally:
            await client.aclose()
        self.assertIsInstance(ctx.exception, DartApiError)
