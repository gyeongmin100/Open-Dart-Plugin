# -*- coding: utf-8 -*-
"""재무제표 모델 JSON -> DART 원문 양식의 Excel 생성.

사용법:
    python build_financial_excel.py --model model.json --output out.xlsx

규칙:
- 재무제표별 시트 + `주석` 시트. 모든 시트 눈금선 숨김.
- 표 범위에만 테두리. 제목/회사명/기간/단위/문단에는 테두리 없음.
- 원문 USERMARK의 글자 크기/서체/굵기, ALIGN/VALIGN 정렬을 반영.
- 긴 텍스트는 wrap + 행 높이 자동 계산으로 항상 전부 보이게.
- 주석번호는 우측 끝 열로 이동, 번호마다 별도 셀 + 파란색/밑줄/하이퍼링크.
- freeze panes, auto filter 사용 금지.
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from _ensure_deps import ensure_deps  # noqa: E402
ensure_deps()

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from dartdoc import inline_note_refs, norm, split_note_refs  # noqa: E402

NOTES_SHEET = "주석"
THIN = Side(style="thin", color="FF808080")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="FFD9D9D9")

MARGIN = 1              # 첫 행/열 여백 (모든 내용은 B2부터)
MARGIN_COL_WIDTH = 2.5
MARGIN_ROW_HEIGHT = 10.0
PARA_COLS = 10          # 주석 문단을 병합할 열 수
NOTES_COL_WIDTH = 13.0  # 주석 시트 기본 열 너비
LINE_HEIGHT = 14.5      # wrap 시 줄당 높이(pt)
DEFAULT_ROW_HEIGHT = 16.5
MAX_ROW_HEIGHT = 409    # Excel 최대 행 높이

_INT_RE = re.compile(r"^-?\d{1,3}(,\d{3})*$|^-?\d+$")
_PAREN_INT_RE = re.compile(r"^\((\d{1,3}(,\d{3})*|\d+)\)$")
# 주석 하위 항목 제목 패턴 (예: "14.1 ...", "가. ...", "(1) ...")
_SUBHEAD_RE = re.compile(r"^(\d{1,3}[.\-]\d|\(?[가-힣]\)?\.\s|\(\d{1,2}\)\s)")


def _sheet_title(name: str, used: set[str]) -> str:
    name = re.sub(r"[\[\]:*?/\\]", " ", name).strip()[:31] or "Sheet"
    base, i = name, 2
    while name in used:
        name = f"{base[:28]}({i})"
        i += 1
    used.add(name)
    return name


def convert_value(text: str):
    """셀 기록 값: 정수 금액은 int, 그 외는 텍스트 그대로. 빈 문자열은 None."""
    stripped = text.strip()
    if _INT_RE.match(stripped):
        return int(stripped.replace(",", ""))
    if _PAREN_INT_RE.match(stripped):
        return -int(stripped[1:-1].replace(",", ""))
    return text if text.strip() != "" else None


def _set_value(cell, text: str):
    value = convert_value(text)
    cell.value = value
    if isinstance(value, int):
        cell.number_format = "#,##0;(#,##0)"


def _disp_len(text: str) -> float:
    """표시 폭(열 너비 단위). 한글/한자/전각은 2, 나머지는 1.1."""
    return sum(2.0 if ord(ch) >= 0x1100 else 1.1 for ch in text)


def _wrap_lines(text: str, width_units: float) -> int:
    """폭 안에 wrap 됐을 때 차지하는 줄 수."""
    usable = max(4.0, width_units - 1.0)
    lines = 0
    for ln in text.split("\n"):
        lines += max(1, math.ceil(_disp_len(ln) / usable))
    return lines


def _bump_row_height(ws, row_idx: int, lines: int, base: float = LINE_HEIGHT):
    if lines <= 1:
        return
    needed = min(MAX_ROW_HEIGHT, lines * base + 4)
    current = ws.row_dimensions[row_idx].height or DEFAULT_ROW_HEIGHT
    if needed > current:
        ws.row_dimensions[row_idx].height = needed


def _font(cell_data: dict | None, *, bold: bool | None = None,
          link: bool = False) -> Font | None:
    """원문 USERMARK 기반 폰트. 지정 없으면 None(기본 폰트)."""
    data = cell_data or {}
    is_bold = data.get("bold", False) if bold is None else bold
    size = data.get("size")
    name = data.get("font")
    if link:
        return Font(name=name, size=size, bold=is_bold,
                    color="FF0000FF", underline="single")
    if not (is_bold or size or name):
        return None
    return Font(name=name, size=size, bold=is_bold)


def _align(cell, horizontal: str | None, wrap: bool = False,
           vertical: str = "center"):
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical,
                               wrap_text=wrap)


def _write_preamble_row(ws, row_idx: int, cells: list[dict], width: int,
                        col_units: list[float] | None = None) -> None:
    """제목/기수/회사명/단위 행 — 테두리 없이 배치."""
    cells = [c for c in cells if c.get("text", "").strip()]
    if not cells:
        return
    total_units = sum(col_units[:width]) if col_units else width * NOTES_COL_WIDTH

    def put(col1: int, col2: int, cell_data: dict, default_align: str):
        col1 += MARGIN
        col2 += MARGIN
        target = ws.cell(row=row_idx, column=col1)
        target.value = cell_data["text"]
        span_units = (sum(col_units[col1 - 1 - MARGIN:col2 - MARGIN])
                      if col_units else (col2 - col1 + 1) * NOTES_COL_WIDTH)
        wrap = _disp_len(cell_data["text"]) > span_units
        _align(target, cell_data.get("align") or default_align, wrap=wrap)
        if wrap:
            _bump_row_height(ws, row_idx,
                             _wrap_lines(cell_data["text"], span_units))
        font = _font(cell_data)
        if font:
            target.font = font
        if col2 > col1:
            ws.merge_cells(start_row=row_idx, start_column=col1,
                           end_row=row_idx, end_column=col2)

    if len(cells) == 1:
        put(1, max(1, width), cells[0], "center")
        return
    # 여러 셀(예: 회사명|단위, 기수 3개 나란히): 폭을 등분해 각 셀 배치
    n = len(cells)
    width = max(width, n)
    bounds = [round(i * width / n) for i in range(n + 1)]
    for i, cell_data in enumerate(cells):
        default = "right" if i == n - 1 and n == 2 else "left"
        put(bounds[i] + 1, bounds[i + 1], cell_data, default)


def _note_tokens(table: dict, note_col: int | None = -1) -> list[list[str]]:
    """행별 주석 토큰. 주석 열이 없으면 빈 리스트들."""
    if note_col == -1:
        note_col = table.get("note_col")
    tokens: list[list[str]] = []
    for row in table["rows"]:
        if note_col is None:
            tokens.append([])
            continue
        cells = row["cells"]
        cell = cells[note_col] if note_col < len(cells) else None
        if row["header"] or cell is None:
            tokens.append([])
        else:
            tokens.append(split_note_refs(cell["text"]))
    return tokens


def _table_col_units(table: dict, note_col: int | None) -> list[float]:
    """출력 열별 너비(열 너비 단위). px WIDTH -> /7 변환."""
    widths = table.get("col_widths") or []
    units = []
    for i, px in enumerate(widths):
        if i == note_col:
            continue
        units.append(min(80.0, max(6.0, px / 7.0)) if px else NOTES_COL_WIDTH)
    return units


def _write_table(ws, start_row: int, table: dict, anchors: dict[str, int],
                 link_stats: dict, relocate_notes: bool = True,
                 col_units: list[float] | None = None) -> int:
    """표를 기록하고 다음 빈 행 번호를 반환. 주석 열은 우측 끝으로 이동.

    relocate_notes=False면 주석 열을 원위치에 그대로 둔다
    (주석 시트 내부 표는 재배치하지 않는다).
    col_units: 출력 열별 너비(wrap/행높이 계산용). 없으면 기본 폭 가정.
    """
    note_col = table.get("note_col") if relocate_notes else None
    grid_width = max((len(r["cells"]) for r in table["rows"]), default=0)
    base_width = grid_width - (1 if note_col is not None else 0)
    tokens_per_row = _note_tokens(table, note_col)
    max_tokens = max((len(t) for t in tokens_per_row), default=0)
    if note_col is not None:
        max_tokens = max(max_tokens, 1)

    # 은행/보험처럼 별도 주석 열 없이 계정명에 "(주석4,6)"이 박힌 경우:
    # 원문 텍스트는 그대로 두고(완전성 보존) 우측에 링크 열을 추가한다.
    inline_mode = False
    if relocate_notes and note_col is None:
        inline_tokens = [
            [] if row["header"] else
            inline_note_refs(" ".join(c["text"] for c in row["cells"] if c))
            for row in table["rows"]
        ]
        if any(inline_tokens):
            inline_mode = True
            tokens_per_row = inline_tokens
            max_tokens = max(len(t) for t in tokens_per_row)

    bordered = not table.get("borderless")

    def out_col(col: int) -> int:
        """원본 열 -> 출력 열(1-based, 여백 열 반영)."""
        if note_col is None or col < note_col:
            return col + 1 + MARGIN
        return col + MARGIN  # col > note_col 은 한 칸 왼쪽으로

    def unit_width(col1: int, col2: int) -> float:
        i1, i2 = col1 - 1 - MARGIN, col2 - MARGIN
        if col_units:
            return sum(col_units[max(0, i1):min(i2, len(col_units))]) or NOTES_COL_WIDTH
        return (col2 - col1 + 1) * NOTES_COL_WIDTH

    # 병합 span 조회용 (wrap 폭 계산에 사용)
    span_of: dict[tuple[int, int], int] = {}
    for r1, c1, r2, c2 in table.get("merges", []):
        span_of[(r1, c1)] = c2 - c1 + 1

    for r, row in enumerate(table["rows"]):
        excel_row = start_row + r
        for c, cell in enumerate(row["cells"]):
            if c == note_col:
                continue
            col_out = out_col(c)
            target = ws.cell(row=excel_row, column=col_out)
            if cell is None:
                continue
            _set_value(target, cell["text"])
            is_header = bool(cell.get("header") or row["header"])
            horizontal = cell.get("align")
            if cell["text"].lstrip().startswith("(단위"):
                horizontal = "right"  # 단위 표시는 항상 우측 정렬
            elif is_header:
                # 열 제목은 가운데 정렬 (원문에 명시 정렬 있으면 존중)
                horizontal = horizontal or "center"
            elif horizontal is None and isinstance(target.value, (int, float)):
                horizontal = "right"
            text = cell["text"]
            span = span_of.get((r, c), 1)
            width_units = unit_width(col_out, col_out + span - 1)
            wrap = "\n" in text or (
                isinstance(target.value, str)
                and _disp_len(text) > width_units)
            _align(target, horizontal, wrap=wrap,
                   vertical=cell.get("valign") or "center")
            if wrap:
                _bump_row_height(ws, excel_row, _wrap_lines(text, width_units))
            font = _font(cell)
            if font:
                target.font = font
            if cell.get("fill"):
                target.fill = PatternFill("solid", fgColor="FF" + cell["fill"])
            elif is_header:
                target.fill = HEADER_FILL  # 열 제목(단위 표시 포함) 회색 배경
        for t, token in enumerate(tokens_per_row[r]):
            target = ws.cell(row=excel_row, column=base_width + 1 + MARGIN + t)
            target.value = token
            _align(target, "center")
            if re.fullmatch(r"\d+", token):
                link_stats["numeric"] += 1
                anchor = anchors.get(token)
                if anchor:
                    # 해당 주석 제목 셀(B열)로 단일 셀 포커스 이동.
                    # 범위 링크는 드래그 선택 상태를 만들어 쓰지 않는다.
                    col = get_column_letter(1 + MARGIN)
                    target.hyperlink = f"#'{NOTES_SHEET}'!{col}{anchor}"
                    target.font = _font(None, link=True)
                    link_stats["linked"] += 1
            else:
                link_stats["other"] += 1

    # 주석 헤더: 재배치된 열(note_col) 또는 인라인 모드로 새로 만든 열에
    # 헤더 블록(다행 헤더 포함) 전체를 덮는 병합 셀로 라벨을 붙인다.
    if note_col is not None or inline_mode:
        header_run = []
        for r, row in enumerate(table["rows"]):
            if not row["header"]:
                break
            header_run.append(r)
        head = None
        if note_col is not None:
            for r in header_run:
                cells = table["rows"][r]["cells"]
                c = cells[note_col] if note_col < len(cells) else None
                if c is not None and c["text"].strip():
                    head = c
                    break
        elif header_run:
            # 인라인 모드는 원문에 전용 헤더가 없으므로 라벨을 새로 만든다
            head = {"text": "주석", "align": None, "bold": True,
                   "size": None, "font": None, "fill": None, "header": True}
        if header_run and head is not None:
            top = start_row + header_run[0]
            target = ws.cell(row=top, column=base_width + 1 + MARGIN)
            target.value = head["text"]
            _align(target, "center")
            target.font = _font(head, bold=True) or Font(bold=True)
            target.fill = (PatternFill("solid", fgColor="FF" + head["fill"])
                           if head.get("fill") else HEADER_FILL)
            bottom = start_row + header_run[-1]
            right = base_width + max_tokens + MARGIN
            if bottom > top or right > base_width + 1 + MARGIN:
                ws.merge_cells(start_row=top,
                               start_column=base_width + 1 + MARGIN,
                               end_row=bottom, end_column=right)

    # 병합 반영 (주석 열은 병합 범위에서 제외하고 클리핑)
    for r1, c1, r2, c2 in table.get("merges", []):
        if note_col is not None:
            if c1 == c2 == note_col:
                continue
            if c1 == note_col:
                c1 += 1
            if c2 == note_col:
                c2 -= 1
        a1, a2 = out_col(c1), out_col(c2)
        if a1 == a2 and r1 == r2:
            continue
        ws.merge_cells(start_row=start_row + r1, start_column=a1,
                       end_row=start_row + r2, end_column=a2)

    # 테두리: 실제 표 범위에만
    if bordered and table["rows"]:
        total_cols = base_width + max_tokens
        for r in range(len(table["rows"])):
            for c in range(1 + MARGIN, total_cols + 1 + MARGIN):
                ws.cell(row=start_row + r, column=c).border = TABLE_BORDER

    return start_row + len(table["rows"])


def _apply_col_widths(ws, table: dict) -> list[float]:
    """열 너비 적용 후 출력 열별 너비 목록 반환."""
    note_col = table.get("note_col")
    units = _table_col_units(table, note_col)
    ws.column_dimensions[get_column_letter(1)].width = MARGIN_COL_WIDTH
    for i, width in enumerate(units):
        ws.column_dimensions[get_column_letter(i + 1 + MARGIN)].width = width
    if note_col is not None:
        ws.column_dimensions[
            get_column_letter(len(units) + 1 + MARGIN)].width = 8
        units = units + [8.0]
    return units


def _write_paragraph(ws, row: int, text: str, *, bold: bool = False,
                     size: int | None = None) -> int:
    cell = ws.cell(row=row, column=1 + MARGIN)
    cell.value = text
    width_units = PARA_COLS * NOTES_COL_WIDTH
    _align(cell, "left", wrap=True, vertical="top")
    if bold or size:
        cell.font = Font(bold=bold, size=size)
    ws.merge_cells(start_row=row, start_column=1 + MARGIN,
                   end_row=row, end_column=PARA_COLS + MARGIN)
    lines = _wrap_lines(text, width_units)
    base = LINE_HEIGHT * (1.15 if size and size > 11 else 1.0)
    ws.row_dimensions[row].height = min(MAX_ROW_HEIGHT, lines * base + 6)
    return row + 1


def _spacer(ws, row: int, height: float = 6.0) -> int:
    ws.row_dimensions[row].height = height
    return row + 1


def _build_notes_sheet(ws, model: dict) -> dict[str, int]:
    """주석 시트를 채우고 {주석번호: 제목 행} 앵커 맵 반환."""
    anchors: dict[str, int] = {}
    ws.row_dimensions[1].height = MARGIN_ROW_HEIGHT
    ws.column_dimensions["A"].width = MARGIN_COL_WIDTH
    row = 1 + MARGIN
    for pre in model.get("notes_preamble", []):
        _write_preamble_row(ws, row, pre, PARA_COLS)
        row += 1
    stats = {"numeric": 0, "linked": 0, "other": 0}

    for note in model.get("notes", []):
        # 주석 사이 큰 여백 (제목 행이 앵커이므로 여백은 제목 앞에)
        row = _spacer(ws, row, 30)
        row = _spacer(ws, row, 30)
        anchors[str(note["number"])] = row
        prev_kind = None
        for i, block in enumerate(note["blocks"]):
            if block["type"] == "paragraph":
                text = block["text"]
                if i == 0:
                    # 주석 제목 행: 굵게 + 크게 (짧을 때만 확대) + 아래 여백
                    row = _write_paragraph(
                        ws, row, text, bold=True,
                        size=13 if len(text) <= 80 else None)
                    row = _spacer(ws, row, 12)
                else:
                    subhead = bool(_SUBHEAD_RE.match(text) and len(text) <= 80)
                    if prev_kind == "table" or subhead:
                        row = _spacer(ws, row, 14)
                    else:
                        row = _spacer(ws, row, 4)
                    row = _write_paragraph(ws, row, text, bold=subhead)
                    if subhead:
                        row = _spacer(ws, row, 6)
                prev_kind = "paragraph"
            else:
                tbl = block["table"]
                cells = [c for r in tbl["rows"] for c in r["cells"]
                         if c and c["text"].strip()]
                is_single_col = all(
                    sum(1 for c in r["cells"] if c and c["text"].strip()) <= 1
                    for r in tbl["rows"])
                # dart4 본문은 서술 문단을 1칸 무테두리 표에 담는다.
                # 빈 칸은 건너뛰고, 내용 칸은 문단으로 렌더링(테두리 없이).
                if tbl.get("borderless") and is_single_col:
                    for c in cells:
                        if prev_kind == "table":
                            row = _spacer(ws, row, 8)
                        row = _write_paragraph(ws, row, c["text"])
                        prev_kind = "paragraph"
                    continue
                if prev_kind == "paragraph":
                    row = _spacer(ws, row, 8)
                row = _write_table(ws, row, tbl, {}, stats,
                                   relocate_notes=False)
                row = _spacer(ws, row, 8)
                prev_kind = "table"

    for col in range(1 + MARGIN, PARA_COLS + 1 + MARGIN):
        ws.column_dimensions[get_column_letter(col)].width = NOTES_COL_WIDTH
    return anchors


def build_workbook(model: dict, output: str) -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    # 앵커를 먼저 확보하기 위해 주석 시트부터 생성
    notes_ws = wb.create_sheet(NOTES_SHEET)
    used.add(NOTES_SHEET)
    anchors = _build_notes_sheet(notes_ws, model)

    link_stats = {"numeric": 0, "linked": 0, "other": 0}
    for statement in model["statements"]:
        ws = wb.create_sheet(_sheet_title(statement["sheet_name"], used))
        first_table = statement["tables"][0] if statement["tables"] else None
        col_units: list[float] = []
        grid_width = 0
        if first_table:
            col_units = _apply_col_widths(ws, first_table)
            grid_width = max(len(r["cells"]) for r in first_table["rows"])
            if first_table.get("note_col") is not None:
                grid_width -= 1
                tokens = _note_tokens(first_table)
                grid_width += max(1, max((len(t) for t in tokens), default=1))
        ws.row_dimensions[1].height = MARGIN_ROW_HEIGHT
        if not col_units:
            ws.column_dimensions["A"].width = MARGIN_COL_WIDTH
        row = 1 + MARGIN
        for pre in statement["preamble"]:
            _write_preamble_row(ws, row, pre, max(1, grid_width),
                                col_units or None)
            # 재무제표 제목 행: 원문에 서식 정보가 없으면 기본 강조
            if (len(pre) == 1 and norm(pre[0]["text"]) == statement["sheet_name"]
                    and not pre[0].get("size")):
                cell = ws.cell(row=row, column=1 + MARGIN)
                cell.font = Font(bold=True, size=14,
                                 name=pre[0].get("font"))
                ws.row_dimensions[row].height = 22
            row += 1
        for table in statement["tables"]:
            row = _write_table(ws, row, table, anchors, link_stats,
                               col_units=col_units or None)
        for text in statement["postscript"]:
            cell = ws.cell(row=row, column=1 + MARGIN)
            cell.value = text
            _align(cell, "left", wrap=False)
            row += 1

    # 눈금선 숨김 (전 시트)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # 주석 시트를 마지막으로 이동
    wb.move_sheet(NOTES_SHEET, offset=len(wb.sheetnames) - 1)
    wb.save(output)
    return {
        "sheets": wb.sheetnames,
        "note_anchors": anchors,
        "links": link_stats,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    model = json.loads(Path(args.model).read_text(encoding="utf-8"))
    result = build_workbook(model, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
