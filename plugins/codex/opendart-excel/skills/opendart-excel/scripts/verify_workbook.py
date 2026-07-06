# -*- coding: utf-8 -*-
"""생성된 Excel 통합문서 검증.

사용법:
    python verify_workbook.py --model model.json --workbook out.xlsx [--source doc.xml]

검증 항목:
- 시트 구성(재무제표 시트들 + `주석`)과 원문 범위 일치
- 재무제표 시트 수(원문 재무제표 수와 동일, 최소 4종 권고)
- 회사명/기간/단위(프리앰블)가 원문과 일치
- 원문에 주석번호가 있으면: 모든 숫자 주석번호가 개별 셀 + 하이퍼링크
- 하이퍼링크가 `주석` 시트의 해당 주석 "제목 행"을 가리키는지
- freeze panes / auto filter 없음
- 주석 번호 연속성(병합/누락 탐지), 참조된 주석이 모두 존재
- --source 지정 시: 원문 재파싱 결과와 모델의 주석 번호 목록 대조

종료코드: 0=통과, 1=실패.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from _ensure_deps import ensure_deps  # noqa: E402
ensure_deps()

from openpyxl import load_workbook  # noqa: E402
import dartdoc  # noqa: E402
from build_financial_excel import convert_value  # noqa: E402
from prepare_notes_json import load_content  # noqa: E402

NOTES_SHEET = "주석"


def _model_note_tokens(model: dict) -> list[str]:
    """모델의 재무제표에서 주석 참조 토큰(원문 기준 기대값).

    별도 주석 열이 있으면 그 열에서, 없으면(은행/보험 등) 계정명에
    박힌 "(주석4,6)" 형태에서 추출한다.
    """
    tokens: list[str] = []
    for statement in model["statements"]:
        for table in statement["tables"]:
            note_col = table.get("note_col")
            for row in table["rows"]:
                if row["header"]:
                    continue
                if note_col is not None:
                    cells = row["cells"]
                    cell = cells[note_col] if note_col < len(cells) else None
                    if cell:
                        tokens.extend(dartdoc.split_note_refs(cell["text"]))
                else:
                    text = " ".join(c["text"] for c in row["cells"] if c)
                    tokens.extend(dartdoc.inline_note_refs(text))
    return tokens


def verify(model: dict, workbook_path: str, source: str | None) -> dict:
    failures: list[str] = []
    warnings: list[str] = []
    wb = load_workbook(workbook_path)

    # 1. 시트 구성
    expected_sheets = []
    used: set[str] = set()
    for s in model["statements"]:
        name = re.sub(r"[\[\]:*?/\\]", " ", s["sheet_name"]).strip()[:31]
        base, i = name, 2
        while name in used:
            name = f"{base[:28]}({i})"
            i += 1
        used.add(name)
        expected_sheets.append(name)
    if wb.sheetnames != expected_sheets + [NOTES_SHEET]:
        failures.append(
            f"시트 구성 불일치: {wb.sheetnames} != {expected_sheets + [NOTES_SHEET]}")

    # 2. 재무제표 수
    n_statements = len(model["statements"])
    if n_statements < 4:
        failures.append(f"재무제표가 {n_statements}종뿐입니다 (최소 4종 기대)")
    elif n_statements == 4:
        warnings.append("재무제표 4종 (손익/포괄손익 통합 형식일 수 있음)")

    # 3. 프리앰블(제목/기간/회사명/단위) 일치
    for s, sheet_name in zip(model["statements"], expected_sheets):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_texts = set()
        for row in ws.iter_rows(max_row=len(s["preamble"]) + 3):
            for cell in row:
                if cell.value is not None:
                    sheet_texts.add(str(cell.value).strip())
        for pre_row in s["preamble"]:
            for cell in pre_row:
                text = cell["text"].strip()
                if text and text not in sheet_texts:
                    failures.append(
                        f"[{sheet_name}] 프리앰블 누락: {text[:50]!r}")

    # 4. 주석 시트 존재
    if NOTES_SHEET not in wb.sheetnames:
        failures.append("`주석` 시트가 없습니다")
        return _report(failures, warnings)
    notes_ws = wb[NOTES_SHEET]

    # 주석 제목 행 맵: 열 A에서 "N." 으로 시작하는 행
    title_rows: dict[str, int] = {}
    for row in notes_ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value is None:
            continue
        m = re.match(r"^(\d{1,3})\.(?!\d)", str(cell.value).strip())
        if m and m.group(1) not in title_rows:
            title_rows[m.group(1)] = cell.row

    # 5. 주석 번호 연속성/누락
    numbers = [n["number"] for n in model["notes"]]
    if len(numbers) != len(set(numbers)):
        failures.append(f"주석 번호 중복: {numbers}")
    if numbers != sorted(numbers):
        failures.append(f"주석 번호가 증가 순서가 아닙니다: {numbers}")
    gaps = [x for a, b in zip(numbers, numbers[1:]) for x in range(a + 1, b)]
    if gaps:
        failures.append(f"주석 번호 누락(병합 의심): {gaps}")
    for n in numbers:
        if str(n) not in title_rows:
            failures.append(f"주석 {n} 제목 행이 `주석` 시트에 없습니다")

    # 6. 주석번호 셀 분리 + 하이퍼링크
    expected_tokens = _model_note_tokens(model)
    expected_numeric = [t for t in expected_tokens if re.fullmatch(r"\d+", t)]
    linkable = [t for t in expected_numeric if int(t) in set(numbers)]
    unlinkable = sorted({t for t in expected_numeric if int(t) not in set(numbers)})
    if unlinkable:
        failures.append(f"재무제표가 참조하는 주석이 파싱되지 않음: {unlinkable}")

    found_links = 0
    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink is None:
                    continue
                target = cell.hyperlink.target or cell.hyperlink.location or ""
                m = re.search(r"'?주석'?![A-Z]+(\d+)", str(target))
                if not m:
                    failures.append(
                        f"[{sheet_name}]{cell.coordinate} 잘못된 링크: {target}")
                    continue
                token = str(cell.value).strip()
                if not re.fullmatch(r"\d+", token):
                    failures.append(
                        f"[{sheet_name}]{cell.coordinate} 주석 셀에 복수 번호/"
                        f"비정상 값: {token!r}")
                    continue
                anchor = int(m.group(1))
                if title_rows.get(token) != anchor:
                    failures.append(
                        f"[{sheet_name}]{cell.coordinate} 주석 {token} 링크가 "
                        f"행 {anchor}를 가리키나 제목 행은 "
                        f"{title_rows.get(token)}입니다")
                fonts_ok = (cell.font and cell.font.underline == "single"
                            and cell.font.color
                            and str(cell.font.color.rgb).endswith("0000FF"))
                if not fonts_ok:
                    failures.append(
                        f"[{sheet_name}]{cell.coordinate} 링크 서식(파랑/밑줄) 아님")
                found_links += 1
    if found_links < len(linkable):
        failures.append(
            f"주석 링크 수 부족: 원문 {len(linkable)}개, 생성 {found_links}개")
    if not expected_numeric:
        warnings.append("원문 재무제표에 주석번호 열이 없음 — 링크 0개는 정상")

    # 6.5 내용 무결성: 모델의 모든 셀 값이 시트에 그대로 존재해야 한다
    def _sheet_counter(ws) -> Counter:
        counter: Counter = Counter()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    counter[cell.value] += 1
        return counter

    def _table_values(table: dict, split_notes: bool = True) -> list:
        values = []
        note_col = table.get("note_col") if split_notes else None
        for row in table["rows"]:
            for c, cell in enumerate(row["cells"]):
                if cell is None:
                    continue
                if c == note_col and not row["header"]:
                    values.extend(dartdoc.split_note_refs(cell["text"]))
                else:
                    v = convert_value(cell["text"])
                    if v is not None:
                        values.append(v)
        return values

    for s, sheet_name in zip(model["statements"], expected_sheets):
        if sheet_name not in wb.sheetnames:
            continue
        expected = Counter()
        for table in s["tables"]:
            expected.update(_table_values(table))
        missing = expected - _sheet_counter(wb[sheet_name])
        if missing:
            sample = list(missing.items())[:5]
            failures.append(
                f"[{sheet_name}] 원문 값이 시트에 없거나 변조됨: {sample}")

    expected_notes = Counter()
    for note in model["notes"]:
        for block in note["blocks"]:
            if block["type"] == "paragraph":
                if block["text"]:
                    expected_notes[block["text"]] += 1
            else:
                # 주석 시트 내부 표는 주석 열을 재배치하지 않으므로 원문 그대로
                expected_notes.update(
                    _table_values(block["table"], split_notes=False))
    missing_notes = expected_notes - _sheet_counter(notes_ws)
    if missing_notes:
        sample = [str(k)[:60] for k in list(missing_notes)[:5]]
        failures.append(f"[주석] 원문 내용이 시트에 없거나 변조됨: {sample}")

    # 주석 제목 행 텍스트가 모델과 일치하는지
    for note in model["notes"]:
        row_idx = title_rows.get(str(note["number"]))
        if row_idx is None:
            continue
        first_block = note["blocks"][0]
        if first_block["type"] != "paragraph":
            continue
        actual = str(notes_ws.cell(row=row_idx, column=2).value or "")
        if actual != first_block["text"]:
            failures.append(
                f"주석 {note['number']} 제목 행 내용이 원문과 다릅니다: "
                f"{actual[:40]!r}")

    # 7. freeze panes / auto filter / 눈금선
    for ws in wb.worksheets:
        if ws.freeze_panes:
            failures.append(f"[{ws.title}] freeze panes 존재: {ws.freeze_panes}")
        if ws.auto_filter.ref:
            failures.append(f"[{ws.title}] auto filter 존재: {ws.auto_filter.ref}")
        if ws.sheet_view.showGridLines:
            failures.append(f"[{ws.title}] 눈금선이 숨겨지지 않음")

    # 8. 원문 재파싱 대조
    if source:
        content = load_content(source)
        fresh = dartdoc.extract_model(content, model["scope"])
        fresh_numbers = [n["number"] for n in fresh["notes"]]
        if fresh_numbers != numbers:
            failures.append(
                f"원문 재파싱 주석 목록 불일치: 원문 {fresh_numbers} != 모델 {numbers}")
        fresh_sheets = [s["sheet_name"] for s in fresh["statements"]]
        model_sheets = [s["sheet_name"] for s in model["statements"]]
        if fresh_sheets != model_sheets:
            failures.append(
                f"원문 재파싱 재무제표 불일치: {fresh_sheets} != {model_sheets}")

        # 9. 글자 단위 완전성: 원문 섹션의 모든 글자가 모델에 있어야 한다
        raw = dartdoc.section_raw_char_counts(content, model["scope"])

        def _chars(texts) -> Counter:
            counter: Counter = Counter()
            for text in texts:
                for ch in text:
                    if not ch.isspace():
                        counter[ch] += 1
            return counter

        def _table_texts(table: dict):
            for row in table["rows"]:
                for cell in row["cells"]:
                    if cell is not None:
                        yield cell["text"]

        st_texts: list[str] = []
        for s in model["statements"]:
            for pre_row in s["preamble"]:
                st_texts.extend(c["text"] for c in pre_row)
            for table in s["tables"]:
                st_texts.extend(_table_texts(table))
            st_texts.extend(s["postscript"])
        nt_texts: list[str] = []
        for pre_row in model.get("notes_preamble", []):
            nt_texts.extend(c["text"] for c in pre_row)
        for note in model["notes"]:
            for block in note["blocks"]:
                if block["type"] == "paragraph":
                    nt_texts.append(block["text"])
                else:
                    nt_texts.extend(_table_texts(block["table"]))

        for label, raw_key in (("재무제표", "statements"), ("주석", "notes")):
            missing = raw[raw_key] - _chars(st_texts if raw_key == "statements"
                                            else nt_texts)
            if missing:
                total = sum(missing.values())
                sample = "".join(ch for ch, _ in missing.most_common(30))
                failures.append(
                    f"[{label}] 원문 글자 {total}개가 결과물에 없음 "
                    f"(예: {sample!r})")

    report = _report(failures, warnings)
    report["stats"] = {
        "sheets": wb.sheetnames,
        "statements": n_statements,
        "notes": len(numbers),
        "expected_note_refs": len(linkable),
        "hyperlinks": found_links,
    }
    return report


def _report(failures: list[str], warnings: list[str]) -> dict:
    return {"ok": not failures, "failures": failures, "warnings": warnings}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", required=True)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--source", help="원문 파일(재파싱 대조용, 권장)")
    args = parser.parse_args()
    model = json.loads(Path(args.model).read_text(encoding="utf-8"))
    report = verify(model, args.workbook, args.source)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    sys.exit(0 if report["ok"] else 1)


if __name__ == "__main__":
    main()
